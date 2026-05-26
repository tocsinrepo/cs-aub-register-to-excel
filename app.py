import streamlit as st
import re
import io
from pdf2image import convert_from_path
import pytesseract
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook.properties import CalcProperties
from datetime import datetime
import tempfile
import os

st.set_page_config(page_title="Bank Register to Excel", layout="wide")
st.title("Bank Register PDF to Excel Converter")
st.write("Upload an Atlantic Union Bank register PDF to convert it into a structured Excel file with debits, credits, and running balances.")

MONTH_MAP = {
    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
    'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DEC': 12
}
MONTH_NAMES = '|'.join(MONTH_MAP.keys())
MONTH_PAT = rf'({MONTH_NAMES})'


def ocr_pdf_to_images(uploaded_file):
    """Convert uploaded PDF to page images."""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(uploaded_file.read())
        tmp_path = tmp.name
    try:
        return convert_from_path(tmp_path, dpi=300)
    finally:
        os.unlink(tmp_path)


def parse_amount(text):
    """Parse a dollar amount string. Parentheses = negative (debit)."""
    if text is None:
        return None
    text = text.strip().rstrip(';:., ')
    neg = re.match(r'^\(\$?([\d,]+\.\d{2})\)$', text)
    if neg:
        return -float(neg.group(1).replace(',', ''))
    pos = re.match(r'^\$?([\d,]+\.\d{2})$', text)
    if pos:
        return float(pos.group(1).replace(',', ''))
    return None


def extract_amounts_from_text(text):
    """Extract all dollar amounts from a text string, returning (amount_str, start_pos) pairs."""
    pattern = r'(\(?\$[\d,]+\.\d{2}\)?)'
    return [(m.group(1), m.start()) for m in re.finditer(pattern, text)]


def is_date_line(line):
    """Check if line is a date like 'FEB 17', a garbled date like 'JAN >', or a year like '2026'."""
    if re.match(rf'^{MONTH_PAT}\s+\d{{1,2}}$', line, re.IGNORECASE):
        return True
    if re.match(rf'^{MONTH_PAT}\s+\S{{1,3}}$', line, re.IGNORECASE):
        return True
    if re.match(r'^[25]\d{3}$', line):
        return True
    return False


def is_amount_line(line):
    """Check if line is a dollar amount (positive or negative in parentheses)."""
    cleaned = line.strip().rstrip(';:., ')
    return bool(re.match(r'^\(?\$?[\d,]+\.\d{2}\)?$', cleaned))


def is_header_or_footer(line):
    """Check if a line is a header, footer, or other non-transaction content."""
    skip_patterns = [
        r'^A(?:tlantic)?$', r'^4?\s*Union Bank', r'^Good Morning',
        r'^Good Afternoon', r'^Good Evening',
        r'^Old Checking', r'^Operations\s+Checking', r'^Operations$',
        r'^Last Updated', r'^Last$', r'^Updated:?$',
        r'^Current Balance', r'^Current$', r'^Available Balance', r'^Available$',
        r'^Transactions\s+Details', r'^Transactions$',
        r'^\$[\d,]+\.\d{2}\s+\$[\d,]+\.\d{2}$',
        r'^Date\s+Description', r'^Amount$', r'^Page totals:',
        r'^\d+\s*-\s*\d+\s+of\s+\d', r'^[<>]+$', r'^Pending\b',
        r'^Details\s*&\s*Settings', r'^Details$', r'^Settings$',
        # Page header keywords appearing together in one line
        r'^.*\bDate\b.*\bBalance\b.*\bDEBIT\b',
        r'^.*\bITEMIZATION\b',
        # Timestamp-only lines (e.g., "3:27 PM")
        r'^\d{1,2}:\d{2}\s*(AM|PM)$',
    ]
    return any(re.match(p, line, re.IGNORECASE) for p in skip_patterns)


# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Positional parser: uses bounding-box data + font-height to correctly match
# each description row with its transaction amount regardless of OCR read order.
# Large-font numbers (h >= 42) = transaction amounts
# Small-font numbers (h < 42)  = running balances
# ---------------------------------------------------------------------------

MIN_TXN_HEIGHT = 42  # pixels at 300 dpi; amounts with h >= this are txn amounts

def parse_page_positional(img, page_num):
    """Parse a page using bounding-box positions and font size (calls image_to_data)."""
    data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
    return parse_page_positional_from_data(data, page_num)


def parse_page_positional_from_data(data, page_num):
    """Parse using pre-computed image_to_data output."""
    elements = []
    for i, text in enumerate(data['text']):
        t = text.strip()
        if not t:
            continue
        elements.append({
            'text': t, 'x': data['left'][i], 'y': data['top'][i],
            'h': data['height'][i], 'w': data['width'][i],
        })

    # --- Identify "Pending" y-positions to exclude pending amounts ---
    pending_y_positions = set()
    for el in elements:
        if el['text'].lower().startswith('pending'):
            pending_y_positions.add(el['y'])
            # Also mark nearby y positions (within 30 pixels)
            for el2 in elements:
                if abs(el2['y'] - el['y']) < 30:
                    pending_y_positions.add(el2['y'])

    # --- Find the y-position of the first month name to separate header from transactions ---
    first_month_y = None
    for el in elements:
        if el['text'].upper() in MONTH_MAP:
            first_month_y = el['y']
            break

    # --- Identify transaction amounts (large font) and running balances (small font) ---
    amt_pattern = re.compile(r'^\(?\$[\d,]+\.\d{2}\)?$')
    txn_amts = []  # large-font dollar amounts
    bal_amts = []  # small-font dollar amounts
    for el in elements:
        if amt_pattern.match(el['text']):
            # Skip amounts in the header area (above the first date) or near "Pending" lines
            if first_month_y is not None and el['y'] < first_month_y - 100:
                continue
            if el['y'] in pending_y_positions:
                continue
            if el['h'] >= MIN_TXN_HEIGHT:
                txn_amts.append(el)
            else:
                bal_amts.append(el)

    txn_amts.sort(key=lambda e: e['y'])
    bal_amts.sort(key=lambda e: e['y'])

    if not txn_amts:
        return []

    # --- Define vertical bands: one per transaction ---
    # Band i runs from midpoint(i-1,i) to midpoint(i,i+1)
    # Cap the first band so it doesn't extend to the top of the page
    # (avoids pulling in header text on page 1)
    MAX_BAND_ABOVE = 150  # max pixels above the amount to look for description
    bands = []
    for i, ta in enumerate(txn_amts):
        if i == 0:
            y_start = max(0, ta['y'] - MAX_BAND_ABOVE)
        else:
            y_start = (txn_amts[i - 1]['y'] + ta['y']) // 2
        y_end = (
            max(el['y'] for el in elements) + 100
            if i == len(txn_amts) - 1
            else (ta['y'] + txn_amts[i + 1]['y']) // 2
        )
        bands.append((y_start, y_end, ta))

    # --- Extract transactions from each band ---
    transactions = []
    for y_start, y_end, txn_el in bands:
        band_els = [el for el in elements if y_start <= el['y'] < y_end]

        month = None
        day = None
        year = 2026
        balance = None
        desc_parts = []

        for el in band_els:
            text = el['text']

            # Skip the transaction amount itself
            if el is txn_el:
                continue

            # Running balance (small-font dollar amount)
            if amt_pattern.match(text) and el['h'] < MIN_TXN_HEIGHT:
                balance = parse_amount(text)
                continue

            # Skip any other dollar amounts
            if amt_pattern.match(text):
                continue

            # Month name
            if text.upper() in MONTH_MAP and month is None:
                month = text.upper()
                continue

            # Year (2026, 5026→2026 etc.)
            if re.match(r'^[25]\d{3}$', text):
                yr = int(text)
                if yr >= 5000:
                    yr -= 3000
                if 2020 <= yr <= 2030:
                    year = yr
                continue

            # Day digit
            if re.match(r'^\d{1,2}$', text):
                d = int(text)
                if 1 <= d <= 31:
                    day = d
                    continue

            # Garbled day (©, >, etc.) next to a month
            if month is not None and day is None and len(text) <= 3:
                fd = fix_garbled_day(text)
                if fd is not None:
                    day = fd
                    continue

            # Skip header/footer and tiny fragments
            if is_header_or_footer(text):
                continue
            if len(text) <= 1 and not text.isalpha():
                continue

            # Description word — keep with x position for ordering
            desc_parts.append((el['x'], text))

        # Build description string left-to-right
        desc_parts.sort(key=lambda p: p[0])
        desc = ' '.join(t for _, t in desc_parts).strip()
        desc = re.sub(r'^[©@&=\-_°®\s]+', '', desc).strip()
        # Remove common OCR garble prefixes before CHECK numbers
        # OCR often prepends garbled text like "5006.", "oe", "Ae", "oon",
        # "Seon", "Noon", "ohne", "otinn", "eS", "o" before CHECK descriptions
        desc = re.sub(r'^[0-9]{4,}\.\s*', '', desc).strip()
        desc = re.sub(r'^,\s*', '', desc).strip()
        # Remove short garbled prefix before CHECK/DEPOSIT/known keywords
        desc = re.sub(r'^[A-Za-z]{1,5}\s+(?=CHECK\b)', '', desc).strip()
        desc = re.sub(r'^[A-Za-z]{1,5}\s+(?=DEPOSIT\b)', '', desc).strip()
        desc = desc.rstrip(':;., ')

        txn_amt = parse_amount(txn_el['text'])

        if txn_amt is not None and month:
            if day is None:
                day = 1
            date_str = f"{MONTH_MAP[month]}/{day}/{year}"
            transactions.append({
                'date': date_str,
                'page': page_num,
                'description': desc or 'UNKNOWN',
                'amount': txn_amt,
                'balance': balance,
            })
        elif txn_amt is not None and month is None and desc:
            # No date in this band — will try to inherit from neighbors below
            transactions.append({
                'date': None,  # placeholder — filled in by date propagation
                'page': page_num,
                'description': desc,
                'amount': txn_amt,
                'balance': balance,
            })

    # --- Date propagation: fill in None dates from nearest dated transaction ---
    # Forward pass: carry last known date downward (older → newer on page)
    last_date = None
    for txn in transactions:
        if txn['date'] is not None:
            last_date = txn['date']
        elif last_date is not None:
            txn['date'] = last_date

    # Backward pass: fill any remaining None dates from below (for bands above the first dated one)
    last_date = None
    for txn in reversed(transactions):
        if txn['date'] is not None:
            last_date = txn['date']
        elif last_date is not None:
            txn['date'] = last_date

    # Remove any transactions that still have no date (header/pending artifacts)
    transactions = [t for t in transactions if t['date'] is not None]

    return transactions


# ---------------------------------------------------------------------------
# Parser for "three-block" pages (standard pages where OCR separates columns)
# OCR output: all dates first, then all descriptions, then all amounts.
# ---------------------------------------------------------------------------

def fix_garbled_day(day_str):
    """Fix OCR-garbled day digits. Returns int or None."""
    OCR_DAY_FIXES = {'>': '5', '<': '1', '|': '1', 'l': '1', 'I': '1',
                     'O': '0', 'o': '0', 'Q': '0', 'D': '0',
                     'S': '5', 's': '5', 'Z': '2', 'z': '2',
                     'B': '8', 'G': '6', 'g': '9', 'q': '9',
                     'T': '7', '?': '9', '!': '1', 'i': '1',
                     'A': '4', 'b': '6', 'e': '8', 'E': '8',
                     '©': '6', '®': '8', '°': '0', ',': '',
                     '.': '', ';': '', ':': '', "'": ''}
    cleaned = ''.join(OCR_DAY_FIXES.get(c, c) for c in day_str)
    cleaned = re.sub(r'[^\d]', '', cleaned)
    if not cleaned:
        return None
    try:
        d = int(cleaned)
        return d if 1 <= d <= 31 else None
    except ValueError:
        return None


def parse_dates_from_raw(dates_raw):
    """Convert raw date lines into 'M/D/YYYY' strings."""
    dates = []
    i = 0
    while i < len(dates_raw):
        m = re.match(rf'^{MONTH_PAT}\s+(\d{{1,2}})$', dates_raw[i], re.IGNORECASE)
        if not m:
            m2 = re.match(rf'^{MONTH_PAT}\s+(\S{{1,3}})$', dates_raw[i], re.IGNORECASE)
            if m2:
                day = fix_garbled_day(m2.group(2))
                if day is not None:
                    month = m2.group(1).upper()
                    yr = 2026
                    if i + 1 < len(dates_raw) and re.match(r'^\d{4}$', dates_raw[i + 1]):
                        yr = int(dates_raw[i + 1])
                        i += 1
                    dates.append(f"{MONTH_MAP[month]}/{day}/{yr}")
            i += 1
            continue
        if m:
            month, day = m.group(1).upper(), int(m.group(2))
            yr = 2026
            if i + 1 < len(dates_raw) and re.match(r'^\d{4}$', dates_raw[i + 1]):
                yr = int(dates_raw[i + 1])
                i += 1
            dates.append(f"{MONTH_MAP[month]}/{day}/{yr}")
        i += 1
    return dates


def pair_amounts(amount_lines):
    """Pair consecutive amount lines into (transaction_amount, running_balance) tuples."""
    pairs = []
    i = 0
    while i < len(amount_lines):
        amt = parse_amount(amount_lines[i])
        bal = None
        if i + 1 < len(amount_lines):
            next_val = parse_amount(amount_lines[i + 1])
            if next_val is not None and next_val > 0:
                bal = next_val
                i += 1
        pairs.append((amt, bal))
        i += 1
    return pairs


def parse_block_page(text, page_num):
    """Parse a page where OCR produces three clean blocks: dates, descriptions, amounts."""
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    filtered = [l for l in lines if not is_header_or_footer(l)]

    dates_raw = []
    descriptions = []
    amount_lines = []
    stray_amounts = []  # (date_count_so_far, amount_string)
    state = 'dates'

    for idx, line in enumerate(filtered):
        if state == 'dates':
            if is_date_line(line):
                dates_raw.append(line)
            elif is_amount_line(line):
                # Check if more dates follow (stray amount in date section)
                lookahead = filtered[idx+1:idx+4]
                dates_ahead = sum(1 for l in lookahead if is_date_line(l))
                if dates_ahead >= 2:
                    # Capture stray amount with current date count for reinsertion
                    stray_amounts.append((len(dates_raw), line.rstrip(';:., ')))
                    continue
                state = 'amounts'
                amount_lines.append(line.rstrip(';:., '))
            else:
                # Check if more dates follow (stray text in date section)
                lookahead = filtered[idx+1:idx+4]
                dates_ahead = sum(1 for l in lookahead if is_date_line(l))
                if dates_ahead >= 2:
                    continue
                state = 'descriptions'
                if len(line) > 2:
                    descriptions.append(line)
        elif state == 'descriptions':
            if is_amount_line(line):
                state = 'amounts'
                amount_lines.append(line.rstrip(';:., '))
            elif not is_date_line(line) and len(line) > 2:
                # Merge continuation lines (previous line ends with \)
                if descriptions and descriptions[-1].rstrip().endswith('\\'):
                    descriptions[-1] = descriptions[-1].rstrip().rstrip('\\').strip() + ' ' + line
                else:
                    descriptions.append(line)
        elif state == 'amounts':
            if is_amount_line(line):
                amount_lines.append(line.rstrip(';:., '))

    # Re-insert stray amounts at the best position (maximise balance chain)
    for stray_date_idx, stray_val in stray_amounts:
        dates_tmp = parse_dates_from_raw(dates_raw)
        best_pos = stray_date_idx * 2        # default: position by date index
        best_score = -1
        # Try every even position (each transaction contributes 2 lines: amount + balance)
        for pos in range(0, len(amount_lines) + 1, 2):
            trial = amount_lines[:pos] + [stray_val] + amount_lines[pos:]
            trial_pairs = pair_amounts(trial)
            tn = min(len(dates_tmp), len(descriptions), len(trial_pairs))
            trial_txns = []
            for ti in range(tn):
                tamt, tbal = trial_pairs[ti]
                if tamt is not None:
                    trial_txns.append({
                        'date': dates_tmp[ti], 'page': page_num,
                        'description': descriptions[ti] if ti < len(descriptions) else '',
                        'amount': tamt, 'balance': tbal,
                    })
            score = validate_balance_chain(trial_txns)
            if score > best_score:
                best_score = score
                best_pos = pos
        amount_lines.insert(best_pos, stray_val)

    dates = parse_dates_from_raw(dates_raw)
    amt_pairs = pair_amounts(amount_lines)

    n = min(len(dates), len(amt_pairs))

    # --- Description alignment: remove excess descriptions guided by balance chain ---
    if len(descriptions) > n:
        excess = len(descriptions) - n
        best_descs = descriptions[:n]
        best_chain = -1
        # Try removing a consecutive block of 'excess' descriptions at each start position
        for start in range(len(descriptions) - excess + 1):
            candidate = descriptions[:start] + descriptions[start + excess:]
            candidate = candidate[:n]
            trial_txns = []
            for i in range(min(n, len(candidate))):
                amt, bal = amt_pairs[i]
                if amt is not None:
                    trial_txns.append({
                        'date': dates[i], 'page': page_num,
                        'description': candidate[i],
                        'amount': amt, 'balance': bal,
                    })
            score = validate_balance_chain(trial_txns)
            if score >= best_chain:
                best_chain = score
                best_descs = candidate[:n]
        descriptions = best_descs
    else:
        descriptions = descriptions[:n]

    transactions = []
    for i in range(n):
        amt, bal = amt_pairs[i]
        if amt is not None:
            transactions.append({
                'date': dates[i],
                'page': page_num,
                'description': descriptions[i] if i < len(descriptions) else '',
                'amount': amt,
                'balance': bal,
            })
    return transactions


def is_block_format(text):
    """Detect if OCR output is in three-block format.

    Tolerates stray amount/text lines among the leading dates.
    If the first ~20 lines are predominantly dates (>60%), it's block format.
    """
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    filtered = [l for l in lines if not is_header_or_footer(l)]

    if len(filtered) < 6:
        return False

    # Check the first 20 filtered lines (or all if fewer)
    check_count = min(20, len(filtered))
    date_count = sum(1 for l in filtered[:check_count] if is_date_line(l))

    # Block format if the leading section is dominated by dates
    return date_count >= 4 and date_count / check_count > 0.6


# ---------------------------------------------------------------------------
# Parser for "merged" pages (first/last pages where OCR mixes columns)
# ---------------------------------------------------------------------------

def parse_merged_page(text, page_num):
    """Parse a page where OCR merges date/description/amount on the same lines."""
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    filtered = [l for l in lines if not is_header_or_footer(l)]

    transactions = []
    i = 0
    while i < len(filtered):
        line = filtered[i]

        date_match = re.match(
            rf'^{MONTH_PAT}\s+(\S{{1,2}})',
            line, re.IGNORECASE
        )
        # Fallback: OCR often omits the space between month and day digit
        # e.g. "APR7  COVA/VENDORPAYM..." or "APR3 MERCHANT BNKCD/..."
        if not date_match:
            date_match = re.match(
                rf'^{MONTH_PAT}(\d{{1,2}})\b',
                line, re.IGNORECASE
            )
        if not date_match:
            i += 1
            continue

        month = date_match.group(1).upper()
        day_str = date_match.group(2).strip()

        OCR_DAY_FIXES = {'>': '5', '<': '1', '|': '1', 'l': '1', 'I': '1',
                         'O': '0', 'o': '0', 'Q': '0', 'D': '0',
                         'S': '5', 's': '5', 'Z': '2', 'z': '2',
                         'B': '8', 'G': '6', 'g': '9', 'q': '9',
                         'T': '7', '?': '9', '!': '1', 'i': '1',
                         'A': '4', 'b': '6', 'e': '8', 'E': '8',
                         '©': '6', '®': '8', '°': '0', ',': '',
                         '.': '', ';': '', ':': '', "'": ''}
        cleaned_day = ''.join(OCR_DAY_FIXES.get(c, c) for c in day_str)
        cleaned_day = re.sub(r'[^\d]', '', cleaned_day)
        if not cleaned_day:
            i += 1
            continue
        try:
            day = int(cleaned_day)
            if day < 1 or day > 31:
                i += 1
                continue
        except ValueError:
            i += 1
            continue

        rest_of_line = line[date_match.end():].strip()

        amounts_in_line = extract_amounts_from_text(rest_of_line)

        if amounts_in_line:
            first_amt_pos = amounts_in_line[0][1]
            desc_part = rest_of_line[:first_amt_pos].strip().rstrip(':;., ')
            txn_amount_str = amounts_in_line[0][0]
        else:
            desc_part = rest_of_line.strip().rstrip(':;., ')
            txn_amount_str = None

        desc_part = re.sub(r'^[_=°®\s]+', '', desc_part).strip()
        # Remove short garbled OCR prefixes before CHECK/DEPOSIT
        desc_part = re.sub(r'^[A-Za-z]{1,5}\s+(?=CHECK\b)', '', desc_part).strip()
        desc_part = re.sub(r'^[A-Za-z]{1,5}\s+(?=DEPOSIT\b)', '', desc_part).strip()
        desc_part = re.sub(r'^,\s*', '', desc_part).strip()

        if not desc_part and i + 1 < len(filtered):
            peek = filtered[i + 1]
            if not re.match(rf'^{MONTH_PAT}\s+', peek, re.IGNORECASE) and \
               not re.match(r'^[25]\d{3}\b', peek) and \
               not is_header_or_footer(peek) and \
               not is_amount_line(peek):
                desc_part = re.sub(r'^[_=°®©@&\s]+', '', peek).strip().rstrip(':;., ')
                i += 1

        year = 2026
        balance = None

        if i + 1 < len(filtered):
            next_line = filtered[i + 1]
            yr_match = re.match(r'^[25]\d{3}\b', next_line)
            if yr_match:
                parsed_yr = int(yr_match.group(0))
                if parsed_yr >= 5000:
                    parsed_yr -= 3000
                if 2020 <= parsed_yr <= 2030:
                    year = parsed_yr

                year_rest = next_line[yr_match.end():].strip()

                amounts_in_year = extract_amounts_from_text(year_rest)

                if amounts_in_year:
                    year_desc = year_rest[:amounts_in_year[0][1]].strip().rstrip(':;., ')
                    year_desc = re.sub(r'^[_=°®\s]+', '', year_desc).strip()

                    if not desc_part and year_desc:
                        desc_part = year_desc
                    elif year_desc and not any(c.isalpha() for c in desc_part):
                        desc_part = year_desc

                    balance = parse_amount(amounts_in_year[-1][0])

                    if txn_amount_str is None and len(amounts_in_year) >= 2:
                        txn_amount_str = amounts_in_year[0][0]
                        balance = parse_amount(amounts_in_year[1][0])
                elif not desc_part:
                    year_desc = year_rest.strip().rstrip(':;., ')
                    year_desc = re.sub(r'^[_=°®\s]+', '', year_desc).strip()
                    if year_desc:
                        desc_part = year_desc

                i += 1

        txn_amount = parse_amount(txn_amount_str) if txn_amount_str else None
        date_str = f"{MONTH_MAP[month]}/{day}/{year}"

        if txn_amount is not None and desc_part:
            transactions.append({
                'date': date_str,
                'page': page_num,
                'description': desc_part,
                'amount': txn_amount,
                'balance': balance,
            })
        elif txn_amount is not None:
            transactions.append({
                'date': date_str,
                'page': page_num,
                'description': 'DEPOSIT',
                'amount': txn_amount,
                'balance': balance,
            })

        i += 1

    return transactions


# ---------------------------------------------------------------------------
# Unified page parser: auto-detect format and dispatch
# ---------------------------------------------------------------------------

def validate_balance_chain(txns):
    """Check how many consecutive transactions have a valid balance chain.

    Transactions are in reverse chronological order (newest first).
    Relationship: balance[i-1] = balance[i] + amount[i-1]
    (the more-recent balance equals the older balance plus the more-recent amount)
    """
    if not txns or len(txns) < 2:
        return len(txns)
    valid = 0
    for i in range(1, len(txns)):
        if txns[i-1]['balance'] is not None and txns[i]['balance'] is not None:
            newer_bal = txns[i - 1]['balance']
            older_bal = txns[i]['balance']
            newer_amt = txns[i - 1]['amount']
            expected = round(older_bal + newer_amt, 2)
            if abs(newer_bal - expected) < 0.02:
                valid += 1
    return valid


def sanitize_phantom_balances(all_txns):
    """Pre-processing: detect OCR phantoms that have fake PDF balances and clear
    those balances so the correction logic can properly handle them.

    Transactions are in reverse chronological order (newest first).
    Balance chain: newer_bal = older_bal + newer_amount.

    Detects three patterns:
    1. Duplicate PDF balance with garbled description.
    2. Balance-equals-amount with garbled description.
    3. Chain-breaking balance: a PDF balance that doesn't fit between its
       neighboring balanced transactions. If skipping it lets the surrounding
       chain work, the balance was likely an OCR misread.
    """
    # --- Pattern 1: Duplicate PDF balances ---
    bal_positions = {}
    for i, t in enumerate(all_txns):
        if t['balance'] is not None:
            bal = round(t['balance'], 2)
            bal_positions.setdefault(bal, []).append(i)

    for bal, indices in bal_positions.items():
        if len(indices) < 2:
            continue
        for idx in indices:
            desc = all_txns[idx].get('description', '')
            alnum_chars = sum(1 for c in desc if c.isalnum())
            has_garbage = any(c in desc for c in '°©®™«»¢£¥')
            if alnum_chars < 5 or has_garbage:
                all_txns[idx]['balance'] = None

    # --- Pattern 2: Balance equals amount + chain inconsistency ---
    # OCR sometimes reads the dollar amount as both the transaction amount AND
    # the running balance. Detect this by checking: if balance ≈ |amount| AND
    # that balance doesn't fit the local chain (large error between this
    # transaction's balance and what the chain predicts), clear the balance.
    bal_indices_p2 = [i for i, t in enumerate(all_txns) if t['balance'] is not None]
    bal_set_p2 = set(bal_indices_p2)
    for pos, curr_i in enumerate(bal_indices_p2):
        t = all_txns[curr_i]
        if abs(t['balance'] - abs(t['amount'])) > 0.50:
            continue  # Not a balance=amount case

        # Also clear if description is garbled (original Pattern 2)
        desc = t.get('description', '')
        alnum_chars = sum(1 for c in desc if c.isalnum())
        has_garbage = any(c in desc for c in '°©®™«»¢£¥')
        if alnum_chars < 5 or has_garbage:
            t['balance'] = None
            continue

        # Check chain fit: does this balance match the previous balanced txn?
        if pos > 0:
            prev_i = bal_indices_p2[pos - 1]
            prev_bal = all_txns[prev_i]['balance']
            # In reverse-chrono: amounts from prev_i to curr_i-1 should sum to
            # prev_bal - curr_bal
            seg_sum = sum(all_txns[k]['amount'] for k in range(prev_i, curr_i))
            expected_diff = round(prev_bal - t['balance'], 2)
            chain_error = abs(seg_sum - expected_diff)
            if chain_error > 100:
                t['balance'] = None
                # Mark as a real transaction with misread balance — do NOT remove
                t['_keep'] = True

    # --- Pattern 3: Chain-breaking balance ---
    # For each balanced transaction, check if its balance fits the local chain.
    # In reverse-chrono order: newer (lower index) → older (higher index).
    # If the chain from prev_balanced to next_balanced works better WITHOUT
    # this transaction's balance, clear it (the balance is likely an OCR misread).
    bal_indices = [i for i, t in enumerate(all_txns) if t['balance'] is not None]
    for pos in range(1, len(bal_indices) - 1):
        prev_i = bal_indices[pos - 1]  # newer balanced transaction
        curr_i = bal_indices[pos]       # current
        next_i = bal_indices[pos + 1]  # older balanced transaction

        prev_bal = all_txns[prev_i]['balance']
        curr_bal = all_txns[curr_i]['balance']
        next_bal = all_txns[next_i]['balance']

        # Check: does curr_bal fit between prev and next?
        # Segment prev→curr: sum of amounts [prev_i .. curr_i-1] should ≈ prev_bal - curr_bal
        seg_prev_curr = sum(all_txns[k]['amount'] for k in range(prev_i, curr_i))
        expected_prev_curr = round(prev_bal - curr_bal, 2)
        error_with = abs(seg_prev_curr - expected_prev_curr)

        # What if we skip curr's balance? Segment prev→next:
        seg_prev_next = sum(all_txns[k]['amount'] for k in range(prev_i, next_i))
        expected_prev_next = round(prev_bal - next_bal, 2)
        error_without = abs(seg_prev_next - expected_prev_next)

        # If skipping this balance dramatically improves chain fit, clear it
        if error_with > 500 and error_without < error_with * 0.25:
            all_txns[curr_i]['balance'] = None
            all_txns[curr_i]['_keep'] = True  # Real txn, just bad balance

    return all_txns


def correct_and_clean_transactions(all_txns):
    """Correct transaction amounts and remove phantoms using the PDF balance chain.

    Transactions are in reverse chronological order (newest first).
    Balance chain: newer_bal = older_bal + newer_amount.

    Approach: find "segments" — spans between two transactions that both have
    PDF balances. Within each segment, the sum of amounts should equal the
    balance difference. If not, try to fix by:
      1. Removing phantom transactions (no PDF balance) that cause the error
      2. Correcting wrong amounts for the remaining transactions

    This handles cases where multiple no-balance transactions sit between
    two balance-carrying transactions.
    """
    if len(all_txns) < 2:
        return all_txns

    # --- Build list of indices that have PDF balances ---
    bal_indices = [i for i, t in enumerate(all_txns) if t['balance'] is not None]

    if len(bal_indices) < 2:
        return all_txns

    # --- Process each segment between consecutive balance-carrying transactions ---
    to_remove = set()

    for seg_idx in range(len(bal_indices) - 1):
        newer_i = bal_indices[seg_idx]      # newer transaction (lower index = more recent)
        older_i = bal_indices[seg_idx + 1]  # older transaction

        newer_bal = all_txns[newer_i]['balance']
        older_bal = all_txns[older_i]['balance']
        expected_sum = round(newer_bal - older_bal, 2)

        # Indices of all transactions in this segment (newer_i to older_i-1 inclusive)
        # These are the transactions whose amounts should sum to expected_sum
        seg_indices = list(range(newer_i, older_i))
        if not seg_indices:
            continue

        actual_sum = round(sum(all_txns[k]['amount'] for k in seg_indices), 2)

        if abs(actual_sum - expected_sum) < 0.02:
            continue  # Segment is already correct

        # --- Try removing no-balance transactions to fix the segment ---
        # Skip transactions marked _keep (real txns with cleared bad balances)
        no_bal_in_seg = [k for k in seg_indices
                         if all_txns[k]['balance'] is None
                         and not all_txns[k].get('_keep')]
        bal_in_seg = [k for k in seg_indices if all_txns[k]['balance'] is not None]

        if no_bal_in_seg:
            # Try each subset of no-balance transactions to remove
            # (for efficiency, try removing one at a time first, then pairs)
            best_remove = None
            best_error = abs(actual_sum - expected_sum)

            # Try removing single no-balance transactions
            for k in no_bal_in_seg:
                trial_sum = round(actual_sum - all_txns[k]['amount'], 2)
                trial_error = abs(trial_sum - expected_sum)
                if trial_error < best_error:
                    best_error = trial_error
                    best_remove = {k}

            # Try removing pairs if singles didn't fix it
            if best_error > 0.02 and len(no_bal_in_seg) >= 2:
                for a_idx in range(len(no_bal_in_seg)):
                    for b_idx in range(a_idx + 1, len(no_bal_in_seg)):
                        k1, k2 = no_bal_in_seg[a_idx], no_bal_in_seg[b_idx]
                        trial_sum = round(
                            actual_sum - all_txns[k1]['amount'] - all_txns[k2]['amount'], 2)
                        trial_error = abs(trial_sum - expected_sum)
                        if trial_error < best_error:
                            best_error = trial_error
                            best_remove = {k1, k2}

            if best_remove and best_error < 0.10:
                to_remove.update(best_remove)
                # Recalculate actual_sum after removal
                actual_sum = round(
                    sum(all_txns[k]['amount'] for k in seg_indices if k not in to_remove), 2)

        # --- Correct amounts for remaining transactions in segment ---
        # Only adjust the head transaction's amount when the remaining segment
        # has NO unverified (no-balance) transactions. If there are no-balance
        # transactions left, their amounts might be wrong, and adjusting the
        # head to compensate would corrupt the head's amount.
        remaining = [k for k in seg_indices if k not in to_remove]
        remaining_sum = round(sum(all_txns[k]['amount'] for k in remaining), 2)
        error = round(remaining_sum - expected_sum, 2)

        if abs(error) > 0.02 and remaining:
            remaining_no_bal = [k for k in remaining if all_txns[k]['balance'] is None]
            if not remaining_no_bal:
                # All remaining transactions have balances — safe to adjust head
                head = newer_i
                if head not in to_remove:
                    others_sum = round(
                        sum(all_txns[k]['amount'] for k in remaining if k != head), 2)
                    correct_head_amt = round(expected_sum - others_sum, 2)
                    all_txns[head]['amount'] = correct_head_amt

    # --- Also correct consecutive pairs as a final pass ---
    for i in range(len(all_txns) - 1):
        if i in to_remove:
            continue
        newer = all_txns[i]
        # Find next non-removed transaction
        older_i = None
        for j in range(i + 1, len(all_txns)):
            if j not in to_remove:
                older_i = j
                break
        if older_i is None:
            continue
        older = all_txns[older_i]

        if newer['balance'] is not None and older['balance'] is not None:
            # Check if there are no transactions between them (after removal)
            between = [k for k in range(i + 1, older_i) if k not in to_remove]
            if not between:
                correct_amt = round(newer['balance'] - older['balance'], 2)
                parsed_amt = round(newer['amount'], 2)
                if abs(correct_amt - parsed_amt) > 0.02:
                    newer['amount'] = correct_amt

    if to_remove:
        return [t for i, t in enumerate(all_txns) if i not in to_remove]
    return all_txns


def validate_global_chain(all_txns):
    """Second-pass validation: walk the full running balance chain and remove
    no-balance transactions whose removal fixes a persistent offset.

    The segment-based correction handles phantoms within segments, but can miss
    phantoms at segment boundaries or cases where the phantom's amount doesn't
    align with a single segment's expected sum. This function uses the global
    running balance (from beginning to end) to detect remaining phantoms.

    Transactions arrive in reverse-chronological order (newest first).
    """
    if len(all_txns) < 2:
        return all_txns

    # Work in chronological order for intuitive chain walking
    txns = list(reversed(all_txns))

    # Find first transaction with a PDF balance to anchor the beginning balance
    first_bal_idx = None
    for i, t in enumerate(txns):
        if t['balance'] is not None:
            first_bal_idx = i
            break
    if first_bal_idx is None:
        return all_txns

    # Compute beginning balance anchored to the first balanced transaction
    amount_sum = sum(t['amount'] for t in txns[:first_bal_idx + 1])
    bb = round(txns[first_bal_idx]['balance'] - amount_sum, 2)

    # Walk the chain looking for offset-causing transactions
    to_remove_chrono = set()
    running = bb

    for i in range(len(txns)):
        if i in to_remove_chrono:
            continue
        running = round(running + txns[i]['amount'], 2)

        if txns[i]['balance'] is not None:
            diff = round(running - txns[i]['balance'], 2)
            if abs(diff) > 0.10:
                # Offset detected. Search backward for a SINGLE no-balance
                # transaction whose amount equals the diff (i.e., removing it
                # would eliminate the offset).
                found = False
                for j in range(i - 1, -1, -1):
                    if j in to_remove_chrono:
                        continue
                    if txns[j]['balance'] is not None:
                        break  # stop at previous balanced transaction
                    if txns[j].get('_keep'):
                        continue  # real txn with cleared balance; don't remove
                    if abs(txns[j]['amount'] - diff) < 0.50:
                        to_remove_chrono.add(j)
                        running = round(running - txns[j]['amount'], 2)
                        found = True
                        break

                # Try pairs if single removal didn't work
                if not found:
                    candidates = []
                    for j in range(i - 1, -1, -1):
                        if j in to_remove_chrono:
                            continue
                        if txns[j]['balance'] is not None:
                            break
                        if txns[j].get('_keep'):
                            continue  # real txn with cleared balance; don't remove
                        candidates.append(j)
                    for a in range(len(candidates)):
                        if found:
                            break
                        for b in range(a + 1, len(candidates)):
                            pair_sum = txns[candidates[a]]['amount'] + txns[candidates[b]]['amount']
                            if abs(pair_sum - diff) < 0.50:
                                to_remove_chrono.add(candidates[a])
                                to_remove_chrono.add(candidates[b])
                                running = round(running - pair_sum, 2)
                                found = True
                                break

    if to_remove_chrono:
        # Convert chronological indices back to reverse-chronological
        n = len(txns)
        to_remove_rev = {n - 1 - j for j in to_remove_chrono}
        return [t for i, t in enumerate(all_txns) if i not in to_remove_rev]
    return all_txns


def remove_garbled_phantoms(all_txns):
    """Final safety net: remove transactions whose description is clearly OCR
    garbage or page header/footer text misread as transactions.

    Detects:
    1. Garbled descriptions (< 5 alnum chars + garbage characters)
    2. Page header text parsed as transactions (contains multiple header keywords)
    3. Timestamp-only descriptions (e.g., "3:27 PM")
    4. Very short garbled descriptions (< 3 alnum chars, no real words)

    Transactions arrive in reverse-chronological order (newest first).
    """
    _garbage_chars = set('°©®™«»¢£¥')
    # Header keywords that indicate page header text, not transactions
    _header_keywords = {'date', 'balance', 'debit', 'credit', 'description',
                        'itemization', 'amount', 'transaction'}
    cleaned = []
    for t in all_txns:
        desc = t.get('description', '')
        desc_lower = desc.lower()
        alnum_count = sum(1 for c in desc if c.isalnum())
        has_garbage = any(c in _garbage_chars for c in desc)

        # Pattern 1: Garbled description with garbage characters
        if alnum_count < 5 and has_garbage:
            continue

        # Pattern 2: Page header text (3+ header keywords in one description)
        header_word_count = sum(1 for kw in _header_keywords if kw in desc_lower)
        if header_word_count >= 3:
            continue

        # Pattern 3: Timestamp-only (e.g., "3:27 PM", "11:05 AM")
        if re.match(r'^\d{1,2}:\d{2}\s*(AM|PM)\s*$', desc, re.IGNORECASE):
            continue

        # Pattern 4: Very short garbled description with no real words
        # (< 5 alnum chars and no recognizable word like CHECK, ACH, ATM)
        if alnum_count < 5 and len(desc) > 0:
            # Check if there's at least one real word (3+ consecutive letters)
            if not re.search(r'[A-Za-z]{3,}', desc):
                continue

        cleaned.append(t)
    return cleaned


def _text_from_data(data):
    """Reconstruct page text from image_to_data() output (grouped by line)."""
    lines = {}
    for i, text in enumerate(data['text']):
        t = text.strip()
        if not t:
            continue
        key = (data['block_num'][i], data['par_num'][i], data['line_num'][i])
        lines.setdefault(key, []).append(t)
    sorted_keys = sorted(lines.keys())
    return '\n'.join(' '.join(lines[k]) for k in sorted_keys)


def parse_page(img, page_num, total_pages):
    """Parse a single page using multiple OCR strategies, keeping the best result."""
    candidates = []

    # --- Default OCR: use image_to_data() to get BOTH text and bounding boxes ---
    # This avoids a separate image_to_string() call (saves ~1.5s per page).
    ocr_data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
    text_default = _text_from_data(ocr_data)
    is_block_default = is_block_format(text_default)
    if is_block_default:
        candidates.append(parse_block_page(text_default, page_num))
    candidates.append(parse_merged_page(text_default, page_num))

    # --- Additional PSM modes (text-only; no bounding boxes needed) ---
    text_psm4 = pytesseract.image_to_string(img, config='--psm 4')
    if is_block_format(text_psm4):
        candidates.append(parse_block_page(text_psm4, page_num))
    candidates.append(parse_merged_page(text_psm4, page_num))

    text_psm6 = pytesseract.image_to_string(img, config='--psm 6')
    if is_block_format(text_psm6):
        candidates.append(parse_block_page(text_psm6, page_num))
    candidates.append(parse_merged_page(text_psm6, page_num))

    # Pick best text-based result
    best = []
    best_score = (-1, -1)
    for c in candidates:
        score = (len(c), validate_balance_chain(c))
        if score > best_score:
            best = c
            best_score = score

    # --- Positional parser (bounding-box + font-size) ---
    # Re-uses the ocr_data already obtained above — no extra OCR call.
    # Matches descriptions to amounts by spatial position, fixing cases
    # where the block parser misaligns them.
    # Always run as a candidate — it handles first/last pages where one date
    # covers multiple transactions (e.g. "APR 22" with 5 checks below it).
    pos = parse_page_positional_from_data(ocr_data, page_num)
    pos_score = (len(pos), validate_balance_chain(pos))
    if pos_score >= best_score:
        best = pos
        best_score = pos_score

    return best


def parse_current_balance(img):
    """Extract the 'Current Balance' value from page 1 header.

    The AUB statement header shows:
        $218,915.24    $1,915.24
        Current Balance  Available Balance

    Returns the current balance as a float, or None if not found.
    """
    w, h = img.size
    header_crop = img.crop((0, 0, w, int(h * 0.35)))
    text = pytesseract.image_to_string(header_crop)
    # Look for dollar amount immediately before "Current Balance"
    match = re.search(r'\$([\d,]+\.\d{2})\s*(?:\$[\d,]+\.\d{2}\s*)?Current\s+Balance',
                      text, re.IGNORECASE)
    if match:
        try:
            return float(match.group(1).replace(',', ''))
        except ValueError:
            pass
    # Fallback: look for the first large dollar amount in the header area
    amounts = re.findall(r'\$([\d,]+\.\d{2})', text)
    if amounts:
        try:
            return float(amounts[0].replace(',', ''))
        except ValueError:
            pass
    return None


def reconcile_with_current_balance(all_txns, current_balance):
    """Check if the transaction list is missing the newest transaction(s)
    by comparing the last PDF balance with the page-1 current balance.

    If there's a gap, insert a synthetic 'MISSING TRANSACTION' entry so the
    balance chain is complete. Transactions are in reverse-chronological order
    (newest first).

    This handles cases where OCR garbles or merges transactions on page 1,
    causing the newest entry (with the highest balance) to be lost.
    """
    if current_balance is None or not all_txns:
        return all_txns

    # The newest transaction is at index 0 (reverse-chrono)
    newest_bal = all_txns[0].get('balance')
    if newest_bal is None:
        # Find first transaction with a balance
        for t in all_txns:
            if t['balance'] is not None:
                newest_bal = t['balance']
                break
    if newest_bal is None:
        return all_txns

    gap = round(current_balance - newest_bal, 2)
    if abs(gap) < 0.50:
        return all_txns  # No gap — balances match

    if gap > 0:
        # Positive gap means there's a missing CREDIT transaction
        # Insert a synthetic transaction at the front (newest position)
        newest_date = all_txns[0].get('date', '')
        synthetic = {
            'date': newest_date,
            'page': 1,
            'description': 'MISSING TRANSACTION (OCR gap — verify against bank statement)',
            'amount': gap,
            'balance': current_balance,
        }
        return [synthetic] + all_txns
    else:
        # Negative gap means there's a missing DEBIT (less common)
        newest_date = all_txns[0].get('date', '')
        synthetic = {
            'date': newest_date,
            'page': 1,
            'description': 'MISSING TRANSACTION (OCR gap — verify against bank statement)',
            'amount': gap,
            'balance': current_balance,
        }
        return [synthetic] + all_txns


def parse_account_info(img):
    """Extract account name from first page header."""
    w, h = img.size
    header_crop = img.crop((0, 0, w, int(h * 0.35)))
    text = pytesseract.image_to_string(header_crop)
    match = re.search(r'((?:Old |New )?Checking Account\s*\*\*\d+)', text)
    return match.group(1).strip() if match else "Bank Register"


# ---------------------------------------------------------------------------
# Read existing Excel register and deduplication
# ---------------------------------------------------------------------------

def read_existing_excel(uploaded_excel):
    """Read an existing Bank Register Excel file and extract transactions + metadata."""
    wb = load_workbook(uploaded_excel, data_only=True)
    ws = wb.active

    # Read beginning balance from row 2, column F (6)
    beginning_balance = ws.cell(row=2, column=6).value or 0

    # Read account name from sheet title or fallback
    account_name = ws.title if ws.title != "Bank Register" else "Bank Register"
    # Try to extract from the file content - check if header exists
    header_val = ws.cell(row=1, column=1).value
    if header_val:
        account_name = "Bank Register"

    transactions = []
    row = 3  # Data starts at row 3 (row 1 = headers, row 2 = beginning balance)
    while row <= ws.max_row:
        date_val = ws.cell(row=row, column=1).value
        if date_val is None or date_val == 'TOTALS' or date_val == 'Total items:':
            break
        # Skip non-transaction rows
        if isinstance(date_val, str) and date_val in ('Beginning Balance', 'TOTALS',
                                                        'Total items:', 'Beginning balance:',
                                                        'Ending balance (Excel):',
                                                        'Ending balance (PDF):',
                                                        'Difference (Excel - PDF):',
                                                        'Reconciliation:'):
            break

        page_val = ws.cell(row=row, column=2).value
        desc_val = ws.cell(row=row, column=3).value or ''
        debit_val = ws.cell(row=row, column=4).value
        credit_val = ws.cell(row=row, column=5).value
        balance_val = ws.cell(row=row, column=6).value
        pdf_bal_val = ws.cell(row=row, column=7).value

        # Reconstruct amount: debits are negative, credits are positive
        if debit_val is not None and debit_val != '' and debit_val != 0:
            amount = -abs(float(debit_val))
        elif credit_val is not None and credit_val != '' and credit_val != 0:
            amount = abs(float(credit_val))
        else:
            row += 1
            continue

        # Handle date - could be string "M/D/YYYY" or datetime object
        if isinstance(date_val, datetime):
            date_str = f"{date_val.month}/{date_val.day}/{date_val.year}"
        else:
            date_str = str(date_val)

        # Handle PDF balance - could be number or 'N/A'
        pdf_balance = None
        if pdf_bal_val is not None and pdf_bal_val != 'N/A':
            try:
                pdf_balance = float(pdf_bal_val)
            except (ValueError, TypeError):
                pdf_balance = None

        transactions.append({
            'date': date_str,
            'page': page_val if page_val else 0,
            'description': desc_val,
            'amount': amount,
            'balance': pdf_balance,
        })
        row += 1

    wb.close()
    return transactions, beginning_balance


def _txn_key(txn):
    """Create a deduplication key from a transaction's date and amount only.

    Description is intentionally excluded because OCR produces different text
    for the same transaction across different PDF scans (e.g. 'CHECK - 10544'
    vs 'CHECK 10544', or garbled prefixes like 'eS CHECK 10508').
    Date + amount is sufficient for deduplication since it's rare to have
    two transactions with the exact same date AND exact same amount.
    """
    amt = round(txn['amount'], 2)
    return (txn['date'], amt)


def deduplicate_transactions(existing_txns, new_txns):
    """Return only transactions from new_txns that don't already exist in existing_txns.

    Uses a multiset approach so that if the same transaction appears twice in the PDF
    and once in Excel, one copy is still added (handles duplicate deposits, etc.).
    Matches by date + amount only (ignores description due to OCR variability).
    """
    # Build a count of existing keys
    existing_counts = {}
    for txn in existing_txns:
        key = _txn_key(txn)
        existing_counts[key] = existing_counts.get(key, 0) + 1

    unique_new = []
    for txn in new_txns:
        key = _txn_key(txn)
        if existing_counts.get(key, 0) > 0:
            existing_counts[key] -= 1  # "consume" one match
        else:
            unique_new.append(txn)

    return unique_new


def parse_date_for_sort(date_str):
    """Parse 'M/D/YYYY' into a sortable datetime."""
    try:
        parts = date_str.split('/')
        return datetime(int(parts[2]), int(parts[0]), int(parts[1]))
    except (ValueError, IndexError):
        return datetime(2000, 1, 1)


def build_excel(transactions, account_name):
    """Build formatted Excel workbook with debits, credits, running balances,
    PDF balances, and a comparison status column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Register"

    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    data_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', bold=True, size=10)
    money_fmt = '#,##0.00'
    thin_border = Border(bottom=Side(style='thin', color='D9D9D9'))

    headers = ['Date', 'Page', 'Description', 'Debits (Out)', 'Credits (In)',
               'Balance', 'PDF Balance', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Reverse to chronological (oldest first)
    transactions.reverse()

    # Beginning balance: anchor from the LAST (most recent) transaction with a
    # PDF balance. This is the most reliable anchor because it's the final
    # statement balance, and it automatically compensates for any phantom
    # transactions that were removed from the middle of the chain.
    #
    # Formula: beginning_balance = last_pdf_balance - sum_of_all_amounts
    last_bal_idx = None
    for i in range(len(transactions) - 1, -1, -1):
        if transactions[i]['balance'] is not None:
            last_bal_idx = i
            break
    if last_bal_idx is not None:
        total_amount_sum = sum(t['amount'] for t in transactions[:last_bal_idx + 1])
        beginning_balance = round(transactions[last_bal_idx]['balance'] - total_amount_sum, 2)
    else:
        # Fallback: try first balanced transaction
        first_bal_idx = None
        for i, txn in enumerate(transactions):
            if txn['balance'] is not None:
                first_bal_idx = i
                break
        if first_bal_idx is not None:
            amount_sum = sum(t['amount'] for t in transactions[:first_bal_idx + 1])
            beginning_balance = round(transactions[first_bal_idx]['balance'] - amount_sum, 2)
        else:
            beginning_balance = 0

    ws.cell(row=2, column=1, value='Beginning Balance').font = bold_font
    ws.cell(row=2, column=6, value=beginning_balance).font = bold_font
    ws.cell(row=2, column=6).number_format = money_fmt

    row = 2
    for txn in transactions:
        row += 1
        ws.cell(row=row, column=1, value=txn['date']).font = data_font
        ws.cell(row=row, column=2, value=txn['page']).font = data_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value=txn['description']).font = data_font

        if txn['amount'] < 0:
            ws.cell(row=row, column=4, value=abs(txn['amount'])).number_format = money_fmt
            ws.cell(row=row, column=4).font = data_font
        else:
            ws.cell(row=row, column=5, value=txn['amount']).number_format = money_fmt
            ws.cell(row=row, column=5).font = data_font

        # Column F: Computed running balance (formula)
        prev = 'F2' if row == 3 else f'F{row-1}'
        ws.cell(row=row, column=6).value = f'={prev}-D{row}+E{row}'
        ws.cell(row=row, column=6).number_format = money_fmt
        ws.cell(row=row, column=6).font = data_font

        # Column G: PDF Balance (from OCR)
        pdf_bal = txn.get('balance')
        if pdf_bal is not None:
            ws.cell(row=row, column=7, value=pdf_bal).number_format = money_fmt
            ws.cell(row=row, column=7).font = data_font
        else:
            ws.cell(row=row, column=7, value='N/A').font = data_font

        # Column H: Status — compare computed balance (F) vs PDF balance (G)
        if pdf_bal is not None:
            ws.cell(row=row, column=8).value = (
                f'=IF(ABS(F{row}-G{row})<0.02,"Match","MISMATCH $"'
                f'&TEXT(F{row}-G{row},"+#,##0.00;-#,##0.00"))'
            )
            ws.cell(row=row, column=8).font = data_font
        else:
            ws.cell(row=row, column=8, value='No PDF bal').font = data_font

        for c in range(1, 9):
            ws.cell(row=row, column=c).border = thin_border

    last_data = row

    # TOTALS row
    row += 1
    ws.cell(row=row, column=1, value='TOTALS').font = bold_font
    fill = PatternFill('solid', fgColor='D9E2F3')
    for c in range(1, 9):
        ws.cell(row=row, column=c).fill = fill
    ws.cell(row=row, column=4).value = f'=SUM(D3:D{last_data})'
    ws.cell(row=row, column=4).number_format = money_fmt
    ws.cell(row=row, column=4).font = bold_font
    ws.cell(row=row, column=5).value = f'=SUM(E3:E{last_data})'
    ws.cell(row=row, column=5).number_format = money_fmt
    ws.cell(row=row, column=5).font = bold_font
    ws.cell(row=row, column=6).value = f'=F{last_data}'
    ws.cell(row=row, column=6).number_format = money_fmt
    ws.cell(row=row, column=6).font = bold_font
    # Count mismatches vs matches
    ws.cell(row=row, column=8).value = (
        f'=COUNTIF(H3:H{last_data},"MISMATCH*")'
        f'&" mismatches / "'
        f'&COUNTIF(H3:H{last_data},"Match")&" matches"'
    )
    ws.cell(row=row, column=8).font = bold_font

    # Summary section
    row += 2
    ws.cell(row=row, column=1, value='Total items:').font = bold_font
    ws.cell(row=row, column=4, value=len(transactions)).font = bold_font

    row += 1
    ws.cell(row=row, column=1, value='Beginning balance:').font = bold_font
    ws.cell(row=row, column=6, value=beginning_balance).font = bold_font
    ws.cell(row=row, column=6).number_format = money_fmt

    row += 1
    ws.cell(row=row, column=1, value='Ending balance (Excel):').font = bold_font
    ws.cell(row=row, column=6).value = f'=F{last_data}'
    ws.cell(row=row, column=6).font = bold_font
    ws.cell(row=row, column=6).number_format = money_fmt

    row += 1
    ws.cell(row=row, column=1, value='Ending balance (PDF):').font = bold_font
    ws.cell(row=row, column=7).value = f'=G{last_data}'
    ws.cell(row=row, column=7).font = bold_font
    ws.cell(row=row, column=7).number_format = money_fmt

    # Difference row
    row += 1
    diff_row = row
    ws.cell(row=row, column=1, value='Difference (Excel - PDF):').font = bold_font
    ws.cell(row=row, column=6).value = f'=F{last_data}-G{last_data}'
    ws.cell(row=row, column=6).font = bold_font
    ws.cell(row=row, column=6).number_format = money_fmt

    # Reconciliation status
    row += 1
    ws.cell(row=row, column=1, value='Reconciliation:').font = bold_font
    ws.cell(row=row, column=6).value = (
        f'=IF(ABS(F{diff_row})<0.02,"BALANCED","OUT OF BALANCE")'
    )
    ws.cell(row=row, column=6).font = bold_font

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 24
    ws.freeze_panes = 'A2'

    # Force Excel/Sheets to recalculate all formulas when file is opened
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

    return wb


# --- Streamlit UI ---
mode = st.radio("Choose mode:", ["New Register", "Update Existing Register"], horizontal=True)

if mode == "New Register":
    st.subheader("Create a New Register from PDF")
    uploaded_file = st.file_uploader("Upload Bank Register PDF", type="pdf", key="new_pdf")

    if uploaded_file:
        if st.button("Convert to Excel"):
            with st.spinner("Running OCR on PDF pages... This may take a minute."):
                images = ocr_pdf_to_images(uploaded_file)
            st.info(f"Processed {len(images)} pages via OCR.")

            with st.spinner("Parsing transactions..."):
                account_name = parse_account_info(images[0])
                all_transactions = []

                for page_num, img in enumerate(images, 1):
                    txns = parse_page(img, page_num, len(images))
                    all_transactions.extend(txns)

            # Remove phantom transactions (OCR artifacts with no PDF balance)
            pre_filter = len(all_transactions)
            # Step 1: Clear fake PDF balances on OCR phantoms
            all_transactions = sanitize_phantom_balances(all_transactions)
            # Step 2: Segment-based correction and phantom removal
            all_transactions = correct_and_clean_transactions(all_transactions)
            # Step 3: Global chain validation to catch remaining phantoms
            all_transactions = validate_global_chain(all_transactions)
            # Step 4: Remove any remaining garbled-description phantoms
            all_transactions = remove_garbled_phantoms(all_transactions)
            removed = pre_filter - len(all_transactions)
            if removed > 0:
                st.info(f"Removed {removed} phantom transaction(s) (OCR artifacts).")

            # Step 5: Reconcile with page-1 current balance
            current_bal = parse_current_balance(images[0])
            if current_bal is not None:
                pre_reconcile = len(all_transactions)
                all_transactions = reconcile_with_current_balance(all_transactions, current_bal)
                if len(all_transactions) > pre_reconcile:
                    gap = round(current_bal - (all_transactions[1].get('balance') or 0), 2)
                    st.warning(f"Detected ${abs(gap):,.2f} gap between last parsed balance and "
                              f"current balance (${current_bal:,.2f}). Added placeholder — "
                              f"verify against your bank statement.")

            st.success(f"Found {len(all_transactions)} transactions from '{account_name}'")

            if all_transactions:
                with_bal = [t for t in all_transactions if t.get('balance') is not None]
                without_bal = len(all_transactions) - len(with_bal)
                st.write(f"**PDF balances found:** {len(with_bal)} of {len(all_transactions)} transactions")
                if without_bal > 0:
                    st.warning(f"{without_bal} transactions have no PDF balance for comparison.")

                preview = []
                for t in all_transactions[:10]:
                    preview.append({
                        'Date': t['date'],
                        'Page': t['page'],
                        'Description': t['description'][:60],
                        'Debit': f"${abs(t['amount']):,.2f}" if t['amount'] < 0 else '',
                        'Credit': f"${t['amount']:,.2f}" if t['amount'] >= 0 else '',
                        'PDF Balance': f"${t['balance']:,.2f}" if t['balance'] else 'N/A',
                    })
                st.write("**Preview (first 10 transactions, newest first):**")
                st.table(preview)

            with st.spinner("Building Excel file..."):
                wb = build_excel(all_transactions, account_name)
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

            st.download_button(
                label="Download Excel File",
                data=output.getvalue(),
                file_name=f"Bank_Register_{account_name.replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.officedocument",
            )

else:
    st.subheader("Update an Existing Register with New PDF Data")
    st.write("Upload your existing Excel register and a new PDF. Only new transactions will be appended.")

    col1, col2 = st.columns(2)
    with col1:
        existing_excel = st.file_uploader("Upload Existing Excel Register", type=["xlsx"], key="existing_xlsx")
    with col2:
        new_pdf = st.file_uploader("Upload New PDF Register", type="pdf", key="update_pdf")

    if existing_excel and new_pdf:
        if st.button("Update Register"):
            # Step 1: Read existing Excel
            with st.spinner("Reading existing Excel register..."):
                existing_txns, existing_beginning_bal = read_existing_excel(existing_excel)
            # Clean OCR phantoms from existing data (may contain artifacts
            # from a previous run that lacked the sanitize fix).
            _garbage_chars = set('°©®™«»¢£¥')
            pre_existing = len(existing_txns)
            existing_txns = [
                t for t in existing_txns
                if not (
                    sum(1 for c in t.get('description', '') if c.isalnum()) < 5
                    and any(c in _garbage_chars for c in t.get('description', ''))
                )
            ]
            existing_removed = pre_existing - len(existing_txns)
            if existing_removed > 0:
                st.info(f"Removed {existing_removed} garbled phantom(s) from existing Excel data.")
            st.info(f"Found {len(existing_txns)} existing transactions in the Excel file.")

            # Step 2: OCR the new PDF
            with st.spinner("Running OCR on new PDF pages... This may take a minute."):
                images = ocr_pdf_to_images(new_pdf)
            st.info(f"Processed {len(images)} pages via OCR.")

            # Step 3: Parse new transactions
            with st.spinner("Parsing new transactions..."):
                account_name = parse_account_info(images[0])
                new_transactions = []
                for page_num, img in enumerate(images, 1):
                    txns = parse_page(img, page_num, len(images))
                    new_transactions.extend(txns)

            # Remove phantom transactions (OCR artifacts with no PDF balance)
            pre_filter = len(new_transactions)
            # Step 1: Clear fake PDF balances on OCR phantoms
            new_transactions = sanitize_phantom_balances(new_transactions)
            # Step 2: Segment-based correction and phantom removal
            new_transactions = correct_and_clean_transactions(new_transactions)
            # Step 3: Global chain validation to catch remaining phantoms
            new_transactions = validate_global_chain(new_transactions)
            # Step 4: Remove any remaining garbled-description phantoms
            new_transactions = remove_garbled_phantoms(new_transactions)
            removed = pre_filter - len(new_transactions)
            if removed > 0:
                st.info(f"Removed {removed} phantom transaction(s) (OCR artifacts).")

            # Step 5: Reconcile with page-1 current balance
            current_bal = parse_current_balance(images[0])
            if current_bal is not None:
                pre_reconcile = len(new_transactions)
                new_transactions = reconcile_with_current_balance(new_transactions, current_bal)
                if len(new_transactions) > pre_reconcile:
                    gap = round(current_bal - (new_transactions[1].get('balance') or 0), 2)
                    st.warning(f"Detected ${abs(gap):,.2f} gap — added placeholder transaction.")

            st.info(f"Found {len(new_transactions)} transactions in the new PDF.")

            # Step 4: Deduplicate
            with st.spinner("Identifying new transactions..."):
                # New PDF transactions come in reverse chronological order (newest first)
                # Reverse them to chronological before deduplication
                new_transactions.reverse()
                unique_new = deduplicate_transactions(existing_txns, new_transactions)

            if not unique_new:
                st.warning("No new transactions found. The existing register is already up to date.")
            else:
                st.success(f"Found **{len(unique_new)}** new transactions to add.")

                # Preview new transactions
                preview = []
                for t in unique_new[:10]:
                    preview.append({
                        'Date': t['date'],
                        'Page': t['page'],
                        'Description': t['description'][:60],
                        'Debit': f"${abs(t['amount']):,.2f}" if t['amount'] < 0 else '',
                        'Credit': f"${t['amount']:,.2f}" if t['amount'] >= 0 else '',
                        'PDF Balance': f"${t['balance']:,.2f}" if t['balance'] else 'N/A',
                    })
                st.write(f"**New transactions to append (showing up to 10 of {len(unique_new)}):**")
                st.table(preview)

                # Step 5: Merge and rebuild
                with st.spinner("Building updated Excel file..."):
                    # Combine existing + new, then sort chronologically
                    all_merged = existing_txns + unique_new
                    all_merged.sort(key=lambda t: parse_date_for_sort(t['date']))

                    # Rebuild the Excel from scratch with all transactions
                    # We need to reverse because build_excel expects newest-first
                    # and will reverse internally to chronological
                    all_merged.reverse()
                    wb = build_excel(all_merged, account_name)
                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)

                # Generate filename based on the uploaded Excel name
                original_name = existing_excel.name
                base_name = original_name.rsplit('.', 1)[0]
                # Increment version number if present (e.g., "04" -> "05")
                ver_match = re.search(r'(\d+)$', base_name)
                if ver_match:
                    old_ver = int(ver_match.group(1))
                    new_ver = str(old_ver + 1).zfill(len(ver_match.group(1)))
                    new_name = base_name[:ver_match.start()] + new_ver + '.xlsx'
                else:
                    new_name = base_name + '_updated.xlsx'

                st.download_button(
                    label="Download Updated Excel File",
                    data=output.getvalue(),
                    file_name=new_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.officedocument",
                )

                st.write(f"**Summary:** {len(existing_txns)} existing + {len(unique_new)} new = {len(existing_txns) + len(unique_new)} total transactions")
